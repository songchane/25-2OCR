import os
import sys
import tempfile
from pathlib import Path
from datetime import datetime
import pandas as pd

from dotenv import load_dotenv
from pipeline import process_documents

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# PDF 출력용
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors


# PDF → 이미지 변환 (PyMuPDF)
def pdf_to_images(pdf_path: Path, dpi: int = 200) -> list[str]:
    try:
        import fitz  # PyMuPDF
    except ImportError:
        print("[오류] pymupdf가 설치되어 있지 않습니다. `pip install pymupdf` 후 다시 시도하세요.")
        sys.exit(1)

    out_dir = Path(tempfile.mkdtemp(prefix="pdf_pages_"))
    image_paths = []
    with fitz.open(pdf_path) as doc:
        for i, page in enumerate(doc):
            pix = page.get_pixmap(dpi=dpi)
            out_path = out_dir / f"{pdf_path.stem}_p{i+1}.png"
            pix.save(out_path.as_posix())
            image_paths.append(out_path.as_posix())
    return image_paths


def collect_inputs(user_path: str) -> list[str]:
    IMG_EXT = {".png", ".jpg", ".jpeg"}
    PDF_EXT = {".pdf"}

    p = Path(user_path)

    if not p.exists():
        print(f"[오류] 입력 경로가 존재하지 않습니다: {user_path}")
        sys.exit(1)

    final_images: list[str] = []

    if p.is_file() and p.suffix.lower() in PDF_EXT:
        final_images.extend(pdf_to_images(p))
    elif p.is_file() and p.suffix.lower() in IMG_EXT:
        final_images.append(p.as_posix())
    elif p.is_dir():
        for f in p.iterdir():
            ext = f.suffix.lower()
            if ext in IMG_EXT:
                final_images.append(f.as_posix())
            elif ext in PDF_EXT:
                final_images.extend(pdf_to_images(f))
    else:
        print("[안내] 처리할 수 있는 PDF나 이미지 파일/폴더가 아닙니다.")
        sys.exit(1)

    if not final_images:
        print("[안내] 입력에서 처리할 파일을 찾지 못했습니다.")
        sys.exit(1)

    return final_images


def save_as_excel(results: dict, out_dir: Path, timestamp: str):
    wb = Workbook()

    # 시트1: 조항 분석
    ws1 = wb.active
    ws1.title = "조항 분석"
    ws1.append(["Clause", "Numbers", "Dates"])
    for c in results.get("clauses", []):
        ws1.append([
            c["clause"],
            ", ".join(c["entities"]["numbers"]),
            ", ".join(c["entities"]["dates"]),
        ])

    d2_font = Font(name="D2Coding", size=10)
    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.font = d2_font

    # 시트2: 안전도 평가
    ws2 = wb.create_sheet("안전도 평가")
    ws2.append(["Clause", "GoodSim", "BadSim", "FinalScore", "Grade"])

    for m in results.get("embedding_match", []):
        ws2.append([
            m["clause"],
            m["good_sim"],
            m["bad_sim"],
            m["final_score"],
            m["grade_text"]
        ])

    # 색상 매핑
    fill_map = {
        "안전": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "경고": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "위험": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
    }

    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.font = d2_font
        grade_value = row[4].value
        if grade_value in fill_map:
            row[4].fill = fill_map[grade_value]

    excel_path = out_dir / f"analysis_result_{timestamp}.xlsx"
    wb.save(excel_path)
    print(f"[저장 완료] Excel → {excel_path}")


def save_as_pdf(results: dict, out_dir: Path, timestamp: str):
    pdf_path = out_dir / f"analysis_result_{timestamp}.pdf"

    font_path = Path("./fonts/D2Coding-Ver1.3.2-20180524.ttf")
    if not font_path.exists():
        raise FileNotFoundError("⚠️ D2Coding 폰트 파일을 ./fonts 폴더에 넣어주세요.")

    pdfmetrics.registerFont(TTFont("D2Coding", str(font_path)))

    styles = getSampleStyleSheet()
    styles["Normal"].fontName = "D2Coding"
    styles["Title"].fontName = "D2Coding"
    styles["Heading2"].fontName = "D2Coding"

    korean_style = ParagraphStyle(
        name="Korean",
        fontName="D2Coding",
        fontSize=9,
        leading=12
    )

    doc = SimpleDocTemplate(str(pdf_path), pagesize=A4)
    elements = []

    elements.append(Paragraph("분석 리포트", styles["Title"]))
    elements.append(Spacer(1, 20))

    # 안전도 평가
    elements.append(Paragraph("안전도 평가", styles["Heading2"]))
    match_data = [[
        Paragraph("Clause", korean_style),
        Paragraph("GoodSim", korean_style),
        Paragraph("BadSim", korean_style),
        Paragraph("FinalScore", korean_style),
        Paragraph("Grade", korean_style)
    ]]

    for m in results.get("embedding_match", []):
        match_data.append([
            Paragraph(m["clause"], korean_style),
            Paragraph(f"{m['good_sim']:.3f}", korean_style),
            Paragraph(f"{m['bad_sim']:.3f}", korean_style),
            Paragraph(f"{m['final_score']:.3f}", korean_style),
            Paragraph(m["grade_text"], korean_style)
        ])

    match_table = Table(match_data, colWidths=[200, 60, 60, 60, 80])
    table_style = TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.grey),
        ("TEXTCOLOR", (0,0), (-1,0), colors.whitesmoke),
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
    ])

    # 색상 처리
    for i, row in enumerate(match_data[1:], start=1):
        grade = row[4].text
        if grade == "안전":
            table_style.add("BACKGROUND", (4, i), (4, i), colors.HexColor("#C6EFCE"))
        elif grade == "경고":
            table_style.add("BACKGROUND", (4, i), (4, i), colors.HexColor("#FFF2CC"))
        elif grade == "위험":
            table_style.add("BACKGROUND", (4, i), (4, i), colors.HexColor("#F8CBAD"))

    match_table.setStyle(table_style)
    elements.append(match_table)

    doc.build(elements)
    print(f"[저장 완료] PDF → {pdf_path}")


def main():
    load_dotenv()
    api_url = os.getenv("API_URL")
    secret_key = os.getenv("SECRET_KEY")
    if not api_url or not secret_key:
        print("[오류] API_URL 또는 SECRET_KEY가 .env에 없습니다.")
        sys.exit(1)

    user_path = input("분석할 이미지 폴더 또는 단일 PDF/이미지 파일 경로를 입력하세요: ").strip()
    user_path = user_path.strip('"').strip("'")
    image_files = collect_inputs(user_path)

    output_format = input("저장 형식을 선택하세요 (excel / pdf / 모두): ").strip().lower()

    # ✅ CSV 로드 (좋은 예시 + 나쁜 예시)
    df = pd.read_csv("D:/25-2ocr2/DB/lease_similarity_pair_db.csv", encoding="utf-8")
    good_examples = df["GoodExample"].dropna().tolist()
    bad_examples = df["BadExample"].dropna().tolist()

    results = process_documents(
        image_files, api_url, secret_key,
        good_examples, bad_examples
    )

    print("=== 최종 분석 결과 ===")
    for r in results["clauses"]:
        print(r)

    print("\n=== 종합 안전도 평가 ===")
    for m in results["embedding_match"]:
        print(m)

    out_dir = Path("D:/25-2ocr2/output")
    out_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if output_format == "excel":
        save_as_excel(results, out_dir, timestamp)
    elif output_format == "pdf":
        save_as_pdf(results, out_dir, timestamp)
    elif output_format == "모두":
        save_as_excel(results, out_dir, timestamp)
        save_as_pdf(results, out_dir, timestamp)
    else:
        print("[안내] 알 수 없는 형식입니다.")


if __name__ == "__main__":
    main()
